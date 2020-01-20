# coding=utf-8
from PyQt5.QtCore import QDir
from openpyxl import Workbook
from openpyxl.styles import Font, colors, Border, Side, Alignment
from openpyxl.worksheet.page import PrintOptions, PrintPageSetup, PageMargins
from datetime import date


class UnionReport(object):
    def __init__(self, ar_karton=None, titan=None, progress=None):
        self.full_massiv = {}
        self.m_result = ['', '', '']

        self._ar_karton = ar_karton
        self._titan = titan

        self.ak = self._titan.__len__()
        self.tl = self._ar_karton.__len__()

        self.len_m = 0
        self.file_name = 'Отчет_{0}.xlsx'.format(date.today())
        self.xlsx_sort = ''

        self.progress = progress

    @property
    def get_name(self):
        return self.file_name

    def Union(self):
        for i in range(self._titan.__len__()):
            self._titan[i][0] = i + 1
            temp = self._titan[i][1]
            self._titan[i][1] = self._titan[i][2]
            self._titan[i][2] = temp
            del self._titan[i][3]
            del self._titan[i][4]
            del self._titan[i][5]

        for i in range(self._titan.__len__()):
            temp = self._titan[i][3]
            self._titan[i][3] = self._titan[i][5]
            self._titan[i][5] = temp

        for i in range(self._ar_karton.__len__()):
            self._ar_karton[i][0] = i + 1
            temp = self._ar_karton[i][1]
            self._ar_karton[i][1] = self._ar_karton[i][2]
            self._ar_karton[i][2] = temp

        if (self.tl - self.ak) > 0:
            self.len_m = self.tl - (abs(self.tl - self.ak))
        else:
            self.len_m = self.ak - (abs(self.tl - self.ak))

        for elem in range(self.len_m - 1):
            if self._titan[elem][2] == self._ar_karton[elem][2]:
                if (self._titan[elem][3] - self._ar_karton[elem][3]) != 0:
                    if (self._titan[elem][3] - self._ar_karton[elem][3]) > 0:
                        self.full_massiv[elem] = self._titan[elem]
                        self.m_result[1] = self._titan[elem][3] - self._ar_karton[elem][3]
                        self.full_massiv[elem] += self.m_result
                        self.full_massiv[elem] += self._ar_karton[elem]
                    else:
                        self.full_massiv[elem] = self._titan[elem]
                        self.m_result[1] = self._titan[elem][3] - self._ar_karton[elem][3]
                        self.full_massiv[elem] += self.m_result
                        self.full_massiv[elem] += self._ar_karton[elem]
                else:
                    continue
            else:
                if self._titan[elem][2] == self._ar_karton[elem + 1][2]:
                    del self._ar_karton[elem]

    def create_xls(self):
        wb = Workbook(write_only=False)
        ws = wb.active
        ws.title = 'Общий отчет'

        r_mass = [[0] * self.full_massiv.keys().__len__()] * 15
        split_list = list(self.full_massiv.values())

        word_wrap_string = Alignment(wrapText=True,
                                     horizontal='center',
                                     vertical='center')

        double_border_side = Side(border_style='dotted')

        square_border = Border(top=double_border_side,
                               right=double_border_side,
                               bottom=double_border_side,
                               left=double_border_side)
        margins_tbls = 0.1968

        ws.page_margins = PageMargins(left=margins_tbls,
                                      right=margins_tbls,
                                      top=margins_tbls,
                                      bottom=margins_tbls)

        ws.sheet_properties.pageSetUpPr.fitToPage = True
        # ws.print_area = 'A1:E13'

        ws.print_options = PrintOptions(horizontalCentered=True)
        ws.page_setup = PrintPageSetup(worksheet=wb.active,
                                       orientation='landscape',
                                       paperSize=ws.PAPERSIZE_A4)

        font_type_tnr = Font(name='Times New Roman', sz=10,
                             color=colors.BLACK, bold=False)
        font_type_arl = Font(name='Arial', sz=10,
                             color=colors.BLACK, bold=True)

        thin = Side(border_style='thin', color='000000')
        range_border = Border(left=thin, right=thin,
                              top=thin, bottom=thin)

        ws.merge_cells('A1:F1')
        ws.merge_cells('J1:O1')

        _headers = {'A1': 'Титан Логистик',
                    'H1': 'Результат',
                    'J1': 'АР Картон'}

        for _key, _values in _headers.items():
            ws[_key].value = _values
            ws[_key].font = font_type_arl
            ws[_key].alignment = word_wrap_string

        for _range in ws.merged_cells.ranges:
            self.style_range(ws, str(_range), border=square_border)

        column_count = {
            'A': 0, 'B': 1, 'C': 2, 'D': 3, 'E': 4,
            'F': 5, 'G': 6, 'H': 7, 'I': 8, 'J': 9,
            'K': 10, 'L': 11, 'M': 12, 'N': 13, 'O': 14
        }

        for c_key, c_values in column_count.items():
            ws.column_dimensions[c_key].width = int(self.max_value(split_list, c_values))

        ws.column_dimensions['G'].width = 0.83
        ws.column_dimensions['I'].width = 0.83
        ws.column_dimensions['H'].width = 11.0

        for i in range(split_list.__len__()):
            for j in range(split_list[i].__len__()):
                ws.cell(i + 2, j + 1).value = split_list[i][j]
                ws.cell(i + 2, j + 1).font = font_type_tnr
            self.progress.setValue(i * 100)

        self.xlsx_sort = '{}{}{}'.format(QDir.homePath(), '/Desktop', '/Общий отчет палетты/')
        if not QDir().exists(self.xlsx_sort):
            QDir().mkdir(self.xlsx_sort)
            wb.save('{}{}'.format(self.xlsx_sort, self.file_name))
            wb.close()
        else:
            wb.save('{}{}'.format(self.xlsx_sort, self.file_name))
            wb.close()

    def max_value(self, array, val_c):
        value_max = len(str(array[1][val_c]))
        for i in range(array.__len__()):
            for j in range(array[i].__len__()):
                if value_max < len(str(array[i][val_c])):
                    value_max = len(str(array[i][val_c]))
                else:
                    continue
        return value_max + 3

    def style_range(self, ws, cell_range, border=Border(), fill=None, font=None, alignment=None):
        top = Border(top=border.top)
        left = Border(left=border.left)
        right = Border(right=border.right)
        bottom = Border(bottom=border.bottom)

        first_cell = ws[cell_range.split(":")[0]]
        if alignment:
            ws.merge_cells(cell_range)
            first_cell.alignment = alignment

        rows = ws[cell_range]
        if font:
            first_cell.font = font

        for cell in rows[0]:
            cell.border = cell.border + top
        for cell in rows[-1]:
            cell.border = cell.border + bottom

        for row in rows:
            l = row[0]
            r = row[-1]
            l.border = l.border + left
            r.border = r.border + right
            if fill:
                for c in row:
                    c.fill = fill

