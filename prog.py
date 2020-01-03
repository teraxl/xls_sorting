from typing import Union

import xlrd
import xlwt
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Side, Border, Font, colors
from openpyxl.worksheet.page import PrintPageSetup, PrintOptions, PageMargins
from openpyxl.worksheet.worksheet import Worksheet


def main():
    wb = Workbook(write_only=False)
    ws = wb.active
    ws.title = 'Общий отчет'

    word_wrap_string = Alignment(wrapText=True,
                                 horizontal='center',
                                 vertical='center')

    double_border_side = Side(border_style='dotted')

    square_border = Border(top=double_border_side,
                           right=double_border_side,
                           bottom=double_border_side,
                           left=double_border_side)
    margins_tbls = 0.1968

    ws.page_margins = PageMargins(left=margins_tbls,
                                  right=margins_tbls,
                                  top=margins_tbls,
                                  bottom=margins_tbls)

    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_area = 'A1:E13'

    ws.print_options = PrintOptions(horizontalCentered=True,
                                    verticalCentered=True)

    ws.page_setup = PrintPageSetup(worksheet=wb.active,
                                   orientation='landscape',
                                   paperSize=ws.PAPERSIZE_A4)

    font_type = Font(name='Times New Roman', sz=14, color=colors.BLACK, bold=True)

    ws.merge_cells('A1:F1')
    ws['A1'] = 'Merge Cells'
    ws['A1'].border = square_border
    ws['A1'].font = font_type
    ws['A1'].alignment = word_wrap_string

    ws.column_dimensions['A'].width = 0.58

    for i in range(2, 13):
        for j in range(1, 15):
            ws.cell(i, j).value = 'text'
            ws.cell(i, j).border = square_border
            ws.cell(i, j).font = font_type
            ws.cell(i, j).alignment = word_wrap_string


    wb.save('000.xlsx')


if __name__ == '__main__':
    main()
